#!/usr/bin/env python3
"""
Export Greenbone/OpenVAS scan reports to a formatted XLSX workbook.

Usage examples:
  python openvas_report_to_xlsx.py --xml report.xml --output report.xlsx

  python openvas_report_to_xlsx.py --gmp-host 192.0.2.10 --username admin \
      --password 'secret' --report-id REPORT_UUID --output report.xlsx
"""

from __future__ import annotations

import argparse
import base64
import html
import re
import socket
import subprocess
import sys
import tempfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


XML_REPORT_FORMAT_ID = "a994b278-1f62-11e1-96ac-406186ea4fc5"


@dataclass
class Finding:
    result_id: str = ""
    host: str = ""
    hostname: str = ""
    port: str = ""
    vulnerability: str = ""
    severity: float = 0.0
    severity_level: str = "Log"
    threat: str = ""
    qod: str = ""
    cves: str = ""
    references: str = ""
    cvss_base: str = ""
    solution_type: str = ""
    summary: str = ""
    impact: str = ""
    affected: str = ""
    detection: str = ""
    insight: str = ""
    solution: str = ""
    raw_description: str = ""


@dataclass
class ReportData:
    report_id: str = ""
    task_name: str = ""
    scan_start: str = ""
    scan_end: str = ""
    timezone: str = ""
    findings: list[Finding] = field(default_factory=list)


def text_of(parent: ET.Element | None, path: str, default: str = "") -> str:
    if parent is None:
        return default
    node = parent.find(path)
    if node is None or node.text is None:
        return default
    return clean_text(node.text)


def clean_text(value: str | None) -> str:
    if not value:
        return ""
    value = html.unescape(value)
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"[ \t]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def severity_level(severity: float, threat: str = "") -> str:
    threat = (threat or "").strip().lower()
    if threat in {"high", "medium", "low", "log", "debug", "false positive"}:
        return threat.title()
    if severity >= 9.0:
        return "Critical"
    if severity >= 7.0:
        return "High"
    if severity >= 4.0:
        return "Medium"
    if severity > 0.0:
        return "Low"
    return "Log"


def parse_tags(tags: str) -> dict[str, str]:
    parsed: dict[str, str] = {}
    if not tags:
        return parsed
    for part in tags.split("|"):
        if "=" not in part:
            continue
        key, value = part.split("=", 1)
        parsed[key.strip().lower()] = clean_text(value)
    return parsed


def parse_float(value: str) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def find_report_root(root: ET.Element) -> ET.Element:
    if root.tag == "report" and root.find("results") is not None:
        return root
    for report in root.iter("report"):
        if report.find("results") is not None:
            return report
    return root


def extract_xml_from_gmp_response(response_xml: bytes) -> bytes:
    root = ET.fromstring(response_xml)
    report = find_report_root(root)
    if report.find("results") is not None:
        return ET.tostring(report, encoding="utf-8")

    # Non-XML formats are returned as base64 in the report element. This path is
    # here only to give a useful error if the wrong report format was requested.
    payload = (report.text or "").strip()
    if payload:
        try:
            return base64.b64decode(payload)
        except Exception:
            pass
    raise ValueError("No XML report payload was found in the GMP response.")


def get_host_value(result: ET.Element) -> tuple[str, str]:
    host_node = result.find("host")
    if host_node is None:
        return "", ""
    host = clean_text(host_node.text)
    hostname = ""
    for detail in host_node.findall("detail"):
        name = text_of(detail, "name").lower()
        value = text_of(detail, "value")
        if name in {"hostname", "host name", "dns"} and value:
            hostname = value
            break
    return host, hostname


def reverse_dns_lookup(host: str, cache: dict[str, str]) -> str:
    if not host:
        return ""
    if host in cache:
        return cache[host]
    try:
        hostname = socket.gethostbyaddr(host)[0].rstrip(".")
    except (OSError, socket.herror, socket.gaierror):
        hostname = ""
    cache[host] = hostname
    return hostname


def refs_from_nvt(nvt: ET.Element | None) -> tuple[str, str]:
    if nvt is None:
        return "", ""
    cves: list[str] = []
    refs: list[str] = []
    for ref in nvt.findall("./refs/ref"):
        ref_type = (ref.get("type") or "").upper()
        ref_id = ref.get("id") or clean_text(ref.text)
        if not ref_id:
            continue
        if ref_type == "CVE":
            cves.append(ref_id)
        refs.append(f"{ref_type}:{ref_id}" if ref_type else ref_id)
    return ", ".join(sorted(set(cves))), ", ".join(sorted(set(refs)))


def parse_report(xml_path: Path) -> ReportData:
    root = ET.parse(xml_path).getroot()
    report = find_report_root(root)
    data = ReportData(
        report_id=report.get("id", ""),
        task_name=text_of(report, "./task/name"),
        scan_start=text_of(report, "scan_start"),
        scan_end=text_of(report, "scan_end"),
        timezone=text_of(report, "timezone"),
    )

    seen: set[str] = set()
    reverse_dns_cache: dict[str, str] = {}
    for result in report.findall("./results/result"):
        result_id = result.get("id", "")
        if result_id and result_id in seen:
            continue
        seen.add(result_id)

        nvt = result.find("nvt")
        tags = parse_tags(text_of(nvt, "tags"))
        host, hostname = get_host_value(result)
        if not hostname:
            hostname = reverse_dns_lookup(host, reverse_dns_cache)
        threat = text_of(result, "threat")
        severity = parse_float(text_of(result, "severity"))
        cves, references = refs_from_nvt(nvt)

        finding = Finding(
            result_id=result_id,
            host=host,
            hostname=hostname,
            port=text_of(result, "port"),
            vulnerability=text_of(result, "name") or text_of(nvt, "name"),
            severity=severity,
            severity_level=severity_level(severity, threat),
            threat=threat,
            qod=text_of(result, "./qod/value"),
            cves=cves,
            references=references,
            cvss_base=text_of(nvt, "cvss_base"),
            solution_type=text_of(nvt, "solution_type"),
            summary=tags.get("summary", ""),
            impact=tags.get("impact", ""),
            affected=tags.get("affected", ""),
            detection=tags.get("vuldetect", ""),
            insight=tags.get("insight", ""),
            solution=text_of(result, "solution") or tags.get("solution", ""),
            raw_description=text_of(result, "description"),
        )
        data.findings.append(finding)

    data.findings.sort(key=lambda item: (item.severity, item.host, item.vulnerability), reverse=True)
    return data


def download_report_with_gvm_cli(args: argparse.Namespace) -> Path:
    filter_arg = args.filter or "apply_overrides=0 levels=hml rows=0 min_qod=70 first=1 sort-reverse=severity"
    request = (
        f'<get_reports report_id="{args.report_id}" '
        f'format_id="{XML_REPORT_FORMAT_ID}" details="1" ignore_pagination="1" '
        f'filter="{html.escape(filter_arg, quote=True)}"/>'
    )
    command = [args.gvm_cli, "--gmp-username", args.username, "--gmp-password", args.password]
    if args.connection == "ssh":
        command.extend(["ssh", "--hostname", args.gmp_host, "--port", str(args.ssh_port), "-A"])
        if args.ssh_username:
            command.extend(["--ssh-username", args.ssh_username])
        if args.ssh_password:
            command.extend(["--ssh-password", args.ssh_password])
    else:
        command.extend(["tls", "--hostname", args.gmp_host, "--port", str(args.gmp_port)])
    command.extend(["--xml", request])
    completed = subprocess.run(command, check=False, capture_output=True, text=False)
    if completed.returncode != 0:
        stdout = completed.stdout.decode("utf-8", errors="replace").strip()
        stderr = completed.stderr.decode("utf-8", errors="replace").strip()
        print("gvm-cli fallo al descargar el reporte.", file=sys.stderr)
        print(f"Codigo de salida: {completed.returncode}", file=sys.stderr)
        if stdout:
            print("", file=sys.stderr)
            print("Salida:", file=sys.stderr)
            print(stdout, file=sys.stderr)
        if stderr:
            print("", file=sys.stderr)
            print("Error:", file=sys.stderr)
            print(stderr, file=sys.stderr)
        print("", file=sys.stderr)
        print("Pistas:", file=sys.stderr)
        print("- En appliances GOS actuales, usa connection=ssh. TLS/9390 puede no estar soportado.", file=sys.stderr)
        print("- Si usas TLS, verifica que el puerto GMP sea el correcto. En instalaciones antiguas suele ser 9390.", file=sys.stderr)
        print("- Verifica usuario, password y que el UUID sea el ID del reporte, no de la tarea.", file=sys.stderr)
        print("- Si usas appliance GOS, confirma que GMP/API este permitido desde tu equipo.", file=sys.stderr)
        raise SystemExit(completed.returncode)
    xml_payload = extract_xml_from_gmp_response(completed.stdout)
    temp_file = Path(tempfile.mkdtemp(prefix="openvas_report_")) / f"{args.report_id}.xml"
    temp_file.write_bytes(xml_payload)
    return temp_file


def add_table(ws, name: str, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    if max_row <= min_row:
        return
    ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def set_widths(ws, widths: dict[str, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def severity_fill(level: str) -> PatternFill:
    colors = {
        "Critical": "C00000",
        "High": "FF0000",
        "Medium": "FFC000",
        "Low": "92D050",
        "Log": "BFBFBF",
        "Debug": "D9D9D9",
        "False Positive": "D9EAD3",
    }
    return PatternFill("solid", fgColor=colors.get(level, "BFBFBF"))


def add_index_sheet(wb: Workbook, data: ReportData) -> None:
    index = wb.create_sheet("Indice", 0)
    index.append(["Hoja", "Contenido"])
    rows = [
        ("Resumen", "Resumen ejecutivo, totales y grafico por severidad."),
        ("Vulnerabilidades", "Todos los hallazgos tecnicos exportados desde OpenVAS."),
        ("Por_Host", "Hallazgos ordenados por activo para remediacion."),
        ("Hosts", "Conteo consolidado por host afectado."),
        ("Vuln_Detalle", "Vulnerabilidades deduplicadas con hosts, impacto y solucion."),
        ("CVEs", "Agrupacion por CVE."),
        ("Remediacion", "Lista priorizada de remediacion."),
    ]
    for sheet, description in rows:
        index.append([sheet, description])
        cell = index.cell(index.max_row, 1)
        cell.hyperlink = f"#'{sheet}'!A1"
        cell.style = "Hyperlink"

    index.append([])
    index.append(["Reporte", data.task_name])
    index.append(["Report ID", data.report_id])
    index.append(["Inicio", data.scan_start])
    index.append(["Fin", data.scan_end])
    style_sheet(index)
    add_table(index, "tblIndice", 1, 1, 8, 2)
    set_widths(index, {"A": 22, "B": 72})


def build_workbook(data: ReportData, output_path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen"

    findings = data.findings
    severity_counts = Counter(item.severity_level for item in findings)
    host_counts = Counter(item.host for item in findings if item.host)
    cve_counts = Counter(
        cve.strip()
        for item in findings
        for cve in item.cves.split(",")
        if cve.strip()
    )

    summary_rows = [
        ["Campo", "Valor"],
        ["Tarea", data.task_name],
        ["Report ID", data.report_id],
        ["Inicio", data.scan_start],
        ["Fin", data.scan_end],
        ["Zona horaria", data.timezone],
        ["Total hallazgos", len(findings)],
        ["Hosts afectados", len(host_counts)],
        ["CVEs únicos", len(cve_counts)],
        ["Severidad máxima", max((f.severity for f in findings), default=0)],
    ]
    for row in summary_rows:
        summary.append(row)

    start = 13
    summary.cell(start, 1, "Severidad")
    summary.cell(start, 2, "Cantidad")
    for offset, level in enumerate(["Critical", "High", "Medium", "Low", "Log"], start=1):
        summary.cell(start + offset, 1, level)
        summary.cell(start + offset, 2, severity_counts.get(level, 0))
        summary.cell(start + offset, 1).fill = severity_fill(level)

    chart = BarChart()
    chart.title = "Hallazgos por severidad"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Severidad"
    values = Reference(summary, min_col=2, min_row=start, max_row=start + 5)
    cats = Reference(summary, min_col=1, min_row=start + 1, max_row=start + 5)
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 12
    summary.add_chart(chart, "D2")
    style_sheet(summary)
    set_widths(summary, {"A": 24, "B": 42, "D": 18, "E": 18})

    vulns = wb.create_sheet("Vulnerabilidades")
    vuln_headers = [
        "Severidad",
        "Nivel",
        "Host",
        "Hostname",
        "Puerto",
        "Vulnerabilidad",
        "QoD",
        "CVEs",
        "Tipo solución",
        "Resumen",
        "Impacto",
        "Solución",
        "Referencias",
        "Result ID",
    ]
    vulns.append(vuln_headers)
    for item in findings:
        vulns.append([
            item.severity,
            item.severity_level,
            item.host,
            item.hostname,
            item.port,
            item.vulnerability,
            item.qod,
            item.cves,
            item.solution_type,
            item.summary,
            item.impact,
            item.solution,
            item.references,
            item.result_id,
        ])
        vulns.cell(vulns.max_row, 2).fill = severity_fill(item.severity_level)
    style_sheet(vulns)
    add_table(vulns, "tblVulnerabilidades", 1, 1, vulns.max_row, len(vuln_headers))
    set_widths(vulns, {
        "A": 11, "B": 13, "C": 16, "D": 24, "E": 18, "F": 46, "G": 8,
        "H": 30, "I": 16, "J": 48, "K": 48, "L": 60, "M": 44, "N": 38,
    })

    by_host: dict[str, list[Finding]] = defaultdict(list)
    for item in findings:
        by_host[item.host].append(item)

    per_host = wb.create_sheet("Por_Host")
    per_host_headers = ["Host", "Hostname", "Severidad", "Nivel", "Puerto", "Vulnerabilidad", "QoD", "CVEs", "Solucion"]
    per_host.append(per_host_headers)
    for host, items in sorted(by_host.items()):
        for item in sorted(items, key=lambda finding: finding.severity, reverse=True):
            per_host.append([
                host,
                item.hostname,
                item.severity,
                item.severity_level,
                item.port,
                item.vulnerability,
                item.qod,
                item.cves,
                item.solution,
            ])
            per_host.cell(per_host.max_row, 4).fill = severity_fill(item.severity_level)
    style_sheet(per_host)
    add_table(per_host, "tblPorHost", 1, 1, per_host.max_row, len(per_host_headers))
    set_widths(per_host, {"A": 18, "B": 28, "C": 11, "D": 13, "E": 18, "F": 52, "G": 8, "H": 30, "I": 80})

    hosts = wb.create_sheet("Hosts")
    hosts.append(["Host", "Hostname", "Total", "Critical", "High", "Medium", "Low", "Log", "Severidad máxima"])
    by_host: dict[str, list[Finding]] = defaultdict(list)
    for item in findings:
        by_host[item.host].append(item)
    for host, items in sorted(by_host.items()):
        host_name = next((item.hostname for item in items if item.hostname), "")
        counts = Counter(item.severity_level for item in items)
        hosts.append([
            host,
            host_name,
            len(items),
            counts.get("Critical", 0),
            counts.get("High", 0),
            counts.get("Medium", 0),
            counts.get("Low", 0),
            counts.get("Log", 0),
            max((item.severity for item in items), default=0),
        ])
    style_sheet(hosts)
    add_table(hosts, "tblHosts", 1, 1, hosts.max_row, 9)
    set_widths(hosts, {"A": 18, "B": 28, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10, "I": 18})

    vuln_detail = wb.create_sheet("Vuln_Detalle")
    detail_headers = [
        "Vulnerabilidad",
        "Severidad maxima",
        "Nivel maximo",
        "Cantidad",
        "Hosts afectados",
        "Puertos",
        "CVEs",
        "Resumen",
        "Impacto",
        "Solucion",
        "Referencias",
    ]
    vuln_detail.append(detail_headers)
    by_vuln: dict[str, list[Finding]] = defaultdict(list)
    for item in findings:
        by_vuln[item.vulnerability].append(item)
    for vulnerability, items in sorted(
        by_vuln.items(),
        key=lambda pair: (max((finding.severity for finding in pair[1]), default=0), len(pair[1])),
        reverse=True,
    ):
        exemplar = max(items, key=lambda finding: finding.severity)
        vuln_detail.append([
            vulnerability,
            exemplar.severity,
            exemplar.severity_level,
            len(items),
            ", ".join(sorted({item.host for item in items if item.host})),
            ", ".join(sorted({item.port for item in items if item.port})),
            ", ".join(sorted({cve.strip() for item in items for cve in item.cves.split(",") if cve.strip()})),
            exemplar.summary,
            exemplar.impact,
            exemplar.solution,
            exemplar.references,
        ])
        vuln_detail.cell(vuln_detail.max_row, 3).fill = severity_fill(exemplar.severity_level)
    style_sheet(vuln_detail)
    add_table(vuln_detail, "tblVulnDetalle", 1, 1, vuln_detail.max_row, len(detail_headers))
    set_widths(vuln_detail, {
        "A": 56, "B": 16, "C": 14, "D": 10, "E": 45, "F": 28, "G": 30,
        "H": 60, "I": 60, "J": 70, "K": 45,
    })

    cves = wb.create_sheet("CVEs")
    cves.append(["CVE", "Cantidad", "Hosts afectados", "Severidad máxima", "Vulnerabilidades"])
    cve_map: dict[str, list[Finding]] = defaultdict(list)
    for item in findings:
        for cve in [part.strip() for part in item.cves.split(",") if part.strip()]:
            cve_map[cve].append(item)
    for cve, items in sorted(cve_map.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        cves.append([
            cve,
            len(items),
            ", ".join(sorted({item.host for item in items if item.host})),
            max((item.severity for item in items), default=0),
            "; ".join(sorted({item.vulnerability for item in items if item.vulnerability})[:8]),
        ])
    style_sheet(cves)
    add_table(cves, "tblCVEs", 1, 1, cves.max_row, 5)
    set_widths(cves, {"A": 18, "B": 10, "C": 45, "D": 18, "E": 80})

    remediation = wb.create_sheet("Remediacion")
    remediation.append(["Prioridad", "Host", "Puerto", "Vulnerabilidad", "Severidad", "Solución"])
    for idx, item in enumerate([f for f in findings if f.severity >= 4.0], start=1):
        remediation.append([idx, item.host, item.port, item.vulnerability, item.severity, item.solution])
    style_sheet(remediation)
    add_table(remediation, "tblRemediacion", 1, 1, remediation.max_row, 6)
    set_widths(remediation, {"A": 10, "B": 18, "C": 18, "D": 52, "E": 12, "F": 80})

    add_index_sheet(wb, data)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 30 if row[0].row > 1 else 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convierte reportes Greenbone/OpenVAS XML a XLSX con hojas de resumen, vulnerabilidades, hosts y CVEs."
    )
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--xml", type=Path, help="Ruta a un reporte XML exportado desde Greenbone.")
    source.add_argument("--gmp-host", help="IP/FQDN del Greenbone Security Assistant/GMP.")
    parser.add_argument("--gmp-port", type=int, default=9390, help="Puerto GMP TLS. Default: 9390.")
    parser.add_argument("--connection", choices=["ssh", "tls"], default="ssh", help="Transporte GMP. Default: ssh.")
    parser.add_argument("--ssh-port", type=int, default=22, help="Puerto SSH para connection=ssh. Default: 22.")
    parser.add_argument("--ssh-username", help="Usuario SSH opcional. Si se omite, gvm-cli usa su default.")
    parser.add_argument("--ssh-password", help="Password SSH opcional. Si se omite, gvm-cli usa su default.")
    parser.add_argument("--username", help="Usuario GMP.")
    parser.add_argument("--password", help="Password GMP.")
    parser.add_argument("--report-id", help="UUID del reporte a descargar.")
    parser.add_argument("--filter", help="Filtro de resultados GMP. Default: niveles H/M/L, QoD >= 70, sin paginación.")
    parser.add_argument("--gvm-cli", default="gvm-cli", help="Ruta al binario gvm-cli. Default: gvm-cli.")
    parser.add_argument("--output", type=Path, default=Path("openvas_report.xlsx"), help="Archivo XLSX de salida.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    if args.gmp_host:
        missing = [name for name in ("username", "password", "report_id") if not getattr(args, name)]
        if missing:
            parser.error("--gmp-host requiere: " + ", ".join(f"--{name.replace('_', '-')}" for name in missing))
        xml_path = download_report_with_gvm_cli(args)
    else:
        xml_path = args.xml

    if not xml_path or not xml_path.exists():
        parser.error(f"No existe el XML: {xml_path}")

    data = parse_report(xml_path)
    build_workbook(data, args.output)
    print(f"OK: {len(data.findings)} hallazgos exportados a {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
